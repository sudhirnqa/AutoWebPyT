import os
import openpyxl

from Utilities import common_utils


def create_excel(file_path):
    """Create a new Excel file."""
    workbook = openpyxl.Workbook()
    workbook.save(file_path)


def read_excel(file_path, sheet_name=''):
    """Read data from an Excel file and return it as a list of dictionaries."""
    log = common_utils.custom_logger()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        data = []
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[i]: row[i] for i in range(len(headers))}
            data.append(row_data)

        log.info(f"Data read from {file_path}: {data}")
        return data

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
        return None
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        log.error(f"Error reading Excel file: {e}")
        return None


def write_excel(file_path, data, sheet_name=''):
    """Write data to an Excel file."""
    log = common_utils.custom_logger()
    try:
        if isinstance(data, list) and all(isinstance(item, dict) for item in data):
            workbook = openpyxl.Workbook()
            sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

            # Write headers
            headers = data[0].keys()
            sheet.append(headers)

            # Write data
            for row in data:
                sheet.append(row.values())

            workbook.save(file_path)

            log.info(f"Data written to {file_path}: {data}")
        else:
            log.error(f"Data must be a list of dictionaries. {data}")
            raise TypeError(f"Data must be a list of dictionaries. {data}")

    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error writing to Excel file: {e}")
        log.error(f"Error writing to Excel file: {e}")


def append_excel(file_path, data, sheet_name=''):
    """Write data to an Excel file."""
    log = common_utils.custom_logger()
    try:
        if isinstance(data, list) and all(isinstance(item, dict) for item in data):
            """Append data to an existing Excel file."""
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
            # Write data
            for row in data:
                sheet.append(row.values())

            workbook.save(file_path)
            log.info(f"Data appended to {file_path}: {data}")
        else:
            log.error(f"Data must be a list of dictionaries. {data}")
            raise TypeError(f"Data must be a list of dictionaries. {data}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error appending to Excel file: {e}")
        log.error(f"Error appending to Excel file: {e}")


def read_excel_column(file_path, column_name, sheet_name=''):
    """Read a specific column from an Excel file and return it as a list."""
    log = common_utils.custom_logger()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        column_data = []

        # Get the index of the specified column
        headers = [cell.value for cell in sheet[1]]
        if column_name in headers:
            column_index = headers.index(column_name) + 1
            for row in sheet.iter_rows(min_row=2, values_only=True):
                column_data.append(row[column_index - 1])
        log.info(f"Data read from {file_path} column {column_name}: {column_data}")
        return column_data
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
        return None
    except Exception as e:
        print(f"Error reading an Excel file: {e}")
        log.error(f"Error reading an Excel file: {e}")
        return None


def write_excel_column(file_path, column_name, data, sheet_name=''):
    """Write data to a specific column in an Excel file."""
    log = common_utils.custom_logger()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

        # Get the index of the specified column
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value=column_name)
            column_index = len(headers) + 1
        else:
            column_index = headers.index(column_name) + 1

        # Write data to the specified column
        for i, value in enumerate(data):
            sheet.cell(row=i + 2, column=column_index, value=value)

        workbook.save(file_path)
        log.info(f"Data written to {file_path} column {column_name}: {data}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error writing an Excel file: {e}")
        log.error(f"Error writing an Excel file: {e}")


def append_excel_column(file_path, column_name, data, sheet_name=''):
    """Append data to a specific column in an Excel file."""
    log = common_utils.custom_logger()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

        # Get the index of the specified column
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            sheet.cell(row=1, column=len(headers) + 1, value=column_name)
            column_index = len(headers) + 1
        else:
            column_index = headers.index(column_name) + 1

        # Append data to the specified column
        for i, value in enumerate(data):
            sheet.cell(row=sheet.max_row + 1, column=column_index, value=value)

        workbook.save(file_path)
        log.info(f"Data appended to {file_path}: {data}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error appending to Excel file: {e}")
        log.error(f"Error appending to Excel file: {e}")


def delete_excel_column(file_path, column_name, sheet_name=''):
    """Delete a specific column from an Excel file."""
    log = common_utils.custom_logger()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

        # Get the index of the specified column
        headers = [cell.value for cell in sheet[1]]
        if column_name in headers:
            column_index = headers.index(column_name) + 1
            sheet.delete_cols(column_index)

        workbook.save(file_path)
        log.info(f"Data deleted in {file_path}: {column_name}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error in deleting a column in Excel file: {e}")
        log.error(f"Error in deleting a column in Excel file: {e}")


def delete_excel_row(file_path, row_number, sheet_name=''):
    """Delete a specific row from an Excel file."""
    log = common_utils.custom_logger()
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

        # Delete the specified row
        sheet.delete_rows(row_number)

        workbook.save(file_path)
        log.info(f"Data deleted in {file_path}: {row_number}")
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        log.error(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error in deleting a row in Excel file: {e}")
        log.error(f"Error in deleting a row in Excel file: {e}")


def delete_file(file_path):
    """Delete a file."""
    log = common_utils.custom_logger()
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            log.info(f"File deleted: {file_path}")
            return True
        return False
    except PermissionError:
        print(f"Permission denied: {file_path}")
        log.error(f"Permission denied: {file_path}")
        return False
    except Exception as e:
        print(f"Error deleting file: {e}")
        log.error(f"Error deleting file: {e}")
        return False


def copy_excel(file_path, new_file_path):
    """Copy an Excel file to a new location."""
    log = common_utils.custom_logger()
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        workbook.save(new_file_path)
        log.info(f"File copied from {file_path} to {new_file_path}")
        return True
    log.error(f"File not found: {file_path}")
    return False


def move_file(file_path, new_file_path):
    """Move a file to a new location."""
    log = common_utils.custom_logger()
    if os.path.exists(file_path):
        os.rename(file_path, new_file_path)
        log.info(f"File moved from {file_path} to {new_file_path}")
        return True
    log.error(f"File not found: {file_path}")
    return False


def rename_file(file_path, new_file_name):
    """Rename a file."""
    log = common_utils.custom_logger()
    if os.path.exists(file_path):
        directory = os.path.dirname(file_path)
        new_file_path = os.path.join(directory, new_file_name)
        os.rename(file_path, new_file_path)
        log.info(f"File renamed from {file_path} to {new_file_path}")
        return True
    log.error(f"File not found: {file_path}")
    return False


def get_file_name(file_path):
    """Get the name of a file without the extension."""
    log = common_utils.custom_logger()
    if os.path.exists(file_path):
        file = os.path.basename(file_path)
        file_name = os.path.splitext(file)[0]
        log.info(f"File name: {file_name}")
        return file_name
    log.error(f"File not found: {file_path}")
    return None

